# Librerías
from extractData.read_pdf import ReadPDF
import os
from openpyxl import Workbook
import json
from exportData.export_to_csv import set_ruta_carpeta

class ExportToExcel():

    # Constructor
    def __init__(self, ruta_carpeta):
        self.ruta_carpeta = ruta_carpeta

        # Obtener las materias desde el archivo JSON
        self.lista_materias = []
        with open('extractData\\subjects.json', 'r') as file:
            self.lista_materias = json.load(file)

        # Crear el archivo Excel
        self.workbook = Workbook()
        self.crear_plantilla_excel()

    ### Métodos ###

    def crear_plantilla_excel(self):
        # Seleccionamos la hoja activa (por defecto, es la primera hoja)
        hoja_activa = self.workbook.active

        # Cambiamos el nombre de la hoja activa
        hoja_activa.title = 'calificaciones'

        # Datos del alumno
        hoja_activa['A1'] = 'apellido_paterno'
        hoja_activa['B1'] = 'apellido_materno'
        hoja_activa['C1'] = 'nombre'
        hoja_activa['D1'] = 'matricula'
        hoja_activa['E1'] = 'periodo'

        # Columnas de materias
        modoPasarMateria = ['_op1', '_op1_extra', '_op2', '_op2_extra', '_op3', '_op3_extra', '_acre', '_reg']
        columna = 6
        for materia, formas in self.lista_materias.items():
            for modo in modoPasarMateria:
                hoja_activa.cell(row=1, column=columna, value=materia + modo)
                columna += 1

    def exportar_a_excel(self):
        # Bandera para saber si existe un PDF en la carpeta
        existsPDF = False

        # Seleccionamos la hoja activa (por defecto, es la primera hoja)
        hoja_activa = self.workbook.active

        # Contador de fila
        contadorFila = 2

        for file in os.listdir(self.ruta_carpeta):

            if file.endswith(".pdf") or file.endswith(".PDF"):
                existsPDF = True

                # Crear la ruta del PDF
                rutaPDF = self.ruta_carpeta + "/" + file
                    
                # Crear un objeto de lectura de PDF
                archivo_pdf = ReadPDF(rutaPDF)

                # Extraer datos del alumno
                nombreCompleto = archivo_pdf.obtener_nombre().split()
                matricula,periodo = archivo_pdf.obtener_matricula_y_periodo()

                # Insertar datos del alumno en la hoja de Excel
                if nombreCompleto:
                    # matricula,periodo=extracion_matricula_periodo(rutaPDF)
                    hoja_activa[f'A{contadorFila}'] = nombreCompleto[len(nombreCompleto) - 2]
                    hoja_activa[f'B{contadorFila}'] = nombreCompleto[len(nombreCompleto) - 1]
                    hoja_activa[f'D{contadorFila}'] = matricula
                    hoja_activa[f'E{contadorFila}'] = periodo

                    if len(nombreCompleto) > 3:
                        hoja_activa[f'C{contadorFila}'] = nombreCompleto[0] + " " + nombreCompleto[1]
                    else:
                        hoja_activa[f'C{contadorFila}'] = nombreCompleto[0]

                # Extraer las calificaciones por materia
                calificacion = []
                columna = 6
                for materia, formas in self.lista_materias.items():
                    for forma in formas:
                        calificacion = archivo_pdf.obtener_calificaciones_materia(forma)

                        if calificacion:
                            # Variable para saber la posición del modo de calificación
                            pos = 0
                            modos = ["ORD", "EXTRA", "ORD", "EXTRA", "ORD", "EXTRA", "ACR", "REG"]

                            for i in range(0, len(calificacion), 2):

                                try:
                                    # Buscar la posición del modo de calificación
                                    while calificacion[i] != modos[pos]:
                                        pos += 1
                                except:
                                    break
                                
                                # Sumarle 1 a la columna para no sobreescribir la materia incorrecta
                                pos += 1

                                # Insertar calificación en la hoja de Excel
                                try:
                                        if calificacion[i+1] == "NP":
                                            hoja_activa.cell(row=contadorFila, column=columna+pos-1, value=(int)(-2))
                                        elif calificacion[i+1] == "SD":
                                            hoja_activa.cell(row=contadorFila, column=columna+pos-1, value=(int)(-3))
                                        else:
                                            hoja_activa.cell(row=contadorFila, column=columna+pos-1, value=(int)(calificacion[i+1]))
                                except:
                                    hoja_activa.cell(row=contadorFila, column=columna+pos-1, value=(int)(-1))
                    
                    # Aumentamos el contador de columnas
                    columna += 8

                # Aumentamos el contador de filas
                contadorFila += 1
                            
        if existsPDF:
            # Ajustar tamaños de las columnas
            for columna in range(1, hoja_activa.max_column + 1):
                letra_columna = hoja_activa.cell(row=1, column=columna).column_letter
                hoja_activa.column_dimensions[letra_columna].width = 30

            # Recorrer las filas y columnas de la hoja para reemplazar los valores nulos por -1
            for fila in hoja_activa.iter_rows():
                for celda in fila:
                    if celda.value is None:
                        celda.value = (int)(-1)

            # Guardamos el archivo Excel
            self.workbook.save(filename=self.ruta_carpeta + '/cardexUABC.xlsx')
            set_ruta_carpeta(self.ruta_carpeta)
            return True
        else:
            set_ruta_carpeta("")
            return False